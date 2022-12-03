import logging
from datetime import datetime
from pygazpar.enum import Frequency
from pygazpar.enum import PropertyName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from typing import Any, List, Dict


FIRST_DATA_LINE_NUMBER = 10

Logger = logging.getLogger(__name__)


# ------------------------------------------------------------------------------------------------------------
class ExcelParser:

    # ------------------------------------------------------
    @staticmethod
    def parse(dataFilename: str, dataReadingFrequency: Frequency) -> List[Dict[str, Any]]:

        parseByFrequency = {
            Frequency.HOURLY: ExcelParser.__parseHourly,
            Frequency.DAILY: ExcelParser.__parseDaily,
            Frequency.WEEKLY: ExcelParser.__parseWeekly,
            Frequency.MONTHLY: ExcelParser.__parseMonthly
        }

        Logger.debug(f"Loading Excel data file '{dataFilename}'...")

        workbook = load_workbook(filename=dataFilename)

        worksheet = workbook.active

        res = parseByFrequency[dataReadingFrequency](worksheet)

        workbook.close()

        return res

    # ------------------------------------------------------
    @staticmethod
    def __fillRow(row: Dict, propertyName: str, cell: Cell, isNumber: bool):

        if cell.value is not None:
            if isNumber:
                if type(cell.value) is str:
                    if len(cell.value.strip()) > 0:
                        row[propertyName] = float(cell.value.replace(',', '.'))
                else:
                    row[propertyName] = cell.value
            else:
                row[propertyName] = cell.value.strip() if type(cell.value) is str else cell.value

    # ------------------------------------------------------
    @staticmethod
    def __parseHourly(worksheet: Worksheet) -> List[Dict[str, Any]]:
        return []

    # ------------------------------------------------------
    @staticmethod
    def __parseDaily(worksheet: Worksheet) -> List[Dict[str, Any]]:

        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()

        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                ExcelParser.__fillRow(row, PropertyName.TIME_PERIOD.value, worksheet.cell(column=2, row=rownum), False)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.START_INDEX.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.END_INDEX.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=5, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=6, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.CONVERTER_FACTOR.value, worksheet.cell(column=7, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.TEMPERATURE.value, worksheet.cell(column=8, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.TYPE.value, worksheet.cell(column=9, row=rownum), False)  # type: ignore
                row[PropertyName.TIMESTAMP.value] = data_timestamp
                res.append(row)

        Logger.debug(f"Daily data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parseWeekly(worksheet: Worksheet) -> List[Dict[str, Any]]:

        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()

        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                ExcelParser.__fillRow(row, PropertyName.TIME_PERIOD.value, worksheet.cell(column=2, row=rownum), False)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                row[PropertyName.TIMESTAMP.value] = data_timestamp
                res.append(row)

        Logger.debug(f"Weekly data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res

    # ------------------------------------------------------
    @staticmethod
    def __parseMonthly(worksheet: Worksheet) -> List[Dict[str, Any]]:

        res = []

        # Timestamp of the data.
        data_timestamp = datetime.now().isoformat()

        minRowNum = FIRST_DATA_LINE_NUMBER
        maxRowNum = len(worksheet['B'])
        for rownum in range(minRowNum, maxRowNum + 1):
            row = {}
            if worksheet.cell(column=2, row=rownum).value is not None:
                ExcelParser.__fillRow(row, PropertyName.TIME_PERIOD.value, worksheet.cell(column=2, row=rownum), False)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.VOLUME.value, worksheet.cell(column=3, row=rownum), True)  # type: ignore
                ExcelParser.__fillRow(row, PropertyName.ENERGY.value, worksheet.cell(column=4, row=rownum), True)  # type: ignore
                row[PropertyName.TIMESTAMP.value] = data_timestamp
                res.append(row)

        Logger.debug(f"Monthly data read successfully between row #{minRowNum} and row #{maxRowNum}")

        return res
