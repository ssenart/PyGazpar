import os
import time
import glob
import logging
from datetime import datetime
from openpyxl import load_workbook
from pygazpar.enum import PropertyNameEnum
from .webdriverwrapper import WebDriverWrapper

HOME_URL = 'https://monespace.grdf.fr'
LOGIN_URL = HOME_URL + '/monespace/connexion'
WELCOME_URL = HOME_URL + '/monespace/particulier/accueil'
DATA_URL = HOME_URL + '/monespace/particulier/consommation/consommations'
DATA_FILENAME = 'Consommations gaz_*.xlsx'

DEFAULT_TMP_DIRECTORY = '/tmp'
DEFAULT_FIREFOX_WEBDRIVER = os.getcwd() + '/geckodriver'
DEFAULT_WAIT_TIME = 30
DEFAULT_LAST_N_ROWS = 0
DEFAULT_HEADLESS_MODE = True


# ------------------------------------------------------------------------------------------------------------
class LoginError(Exception):
    """ Client has failed to login in GrDF Web site (check username/password)"""
    pass


# ------------------------------------------------------------------------------------------------------------
class Client(object):

    logger = logging.getLogger(__name__)

    # ------------------------------------------------------
    def __init__(self, username: str, password: str, firefox_webdriver_executable: str = DEFAULT_FIREFOX_WEBDRIVER, wait_time: int = DEFAULT_WAIT_TIME, tmp_directory: str = DEFAULT_TMP_DIRECTORY, lastNRows: int = DEFAULT_LAST_N_ROWS, headLessMode: bool = DEFAULT_HEADLESS_MODE):
        self.__username = username
        self.__password = password
        self.__firefox_webdriver_executable = firefox_webdriver_executable
        self.__wait_time = wait_time
        self.__tmp_directory = tmp_directory
        self.__data = []
        self.__lastNRows = lastNRows
        self.__headlessMode = headLessMode

    # ------------------------------------------------------
    def data(self):
        return self.__data

    # ------------------------------------------------------
    def acceptCookies(self, driver: WebDriverWrapper):

        try:
            cookies_accept_button = driver.find_element_by_xpath("//a[@id='_EPcommonPage_WAR_EPportlet_:formBandeauCnil:j_idt12']", "Cookies accept button", False)
            cookies_accept_button.click()
        except Exception:
            # Do nothing, because the Pop up may not appear.
            pass

    # ------------------------------------------------------
    def acceptPrivacyConditions(self, driver: WebDriverWrapper):

        try:
            # id=btn_accept_banner
            accept_button = driver.find_element_by_id("btn_accept_banner", "Privacy Conditions accept button", False)
            accept_button.click()
        except Exception:
            # Do nothing, because the Pop up may not appear.
            pass

    # ------------------------------------------------------
    def closeEventualPopup(self, driver: WebDriverWrapper):

        # Accept an eventual Privacy Conditions popup.
        self.acceptPrivacyConditions(driver)

        # Eventually, click Accept in the lower banner to accept cookies from the site.
        self.acceptCookies(driver)

        # Eventually, close Advertisement Popup Windows.
        try:
            advertisement_popup_element = driver.find_element_by_xpath("/html/body/abtasty-modal/div/div[1]", "Advertisement close button", False)
            advertisement_popup_element.click()
        except Exception:
            # Do nothing, because the Pop up may not appear.
            pass

        # Eventually, close Survey Popup Windows : /html/body/div[12]/div[2] or //*[@id="mfbIframeClose"]
        try:
            survey_popup_element = driver.find_element_by_xpath("//*[@id='mfbIframeClose']", "Survey close button", False)
            survey_popup_element.click()
        except Exception:
            # Do nothing, because the Pop up may not appear.
            pass

    # ------------------------------------------------------
    def update(self):

        Client.logger.debug("Start updating the data...")

        # XLSX is in the TMP directory
        data_file_path_pattern = self.__tmp_directory + '/' + DATA_FILENAME

        # We remove an eventual existing data file (from a previous run that has not deleted it)
        file_list = glob.glob(data_file_path_pattern)
        for filename in file_list:
            if os.path.isfile(filename):
                os.remove(filename)

        # Create the WebDriver with the ability to log and take screenshot for debugging.
        driver = WebDriverWrapper(self.__firefox_webdriver_executable, self.__wait_time, self.__tmp_directory, self.__headlessMode)

        try:

            # Login URL
            driver.get(LOGIN_URL, "Go to login page")

            # Fill login form
            email_element = driver.find_element_by_id("_EspacePerso_WAR_EPportlet_:seConnecterForm:email", "Login page: Email text field")
            password_element = driver.find_element_by_id("_EspacePerso_WAR_EPportlet_:seConnecterForm:passwordSecretSeConnecter", "Login page: Password text field")

            email_element.send_keys(self.__username)
            password_element.send_keys(self.__password)

            # Submit the login form.
            submit_button_element = driver.find_element_by_id("_EspacePerso_WAR_EPportlet_:seConnecterForm:meConnecter", "Login page: 'Me connecter' button")
            submit_button_element.click()

            # Close eventual popup Windows or Assistant appearing.
            self.closeEventualPopup(driver)

            # Once we find the 'Acceder' button from the main page, we are logged on successfully.
            try:
                driver.find_element_by_xpath("//div[2]/div[2]/div/a/div", "Welcome page: 'Acceder' button of 'Suivi de consommation'")
            except Exception:
                # Perhaps, login has failed.
                if driver.current_url() == WELCOME_URL:
                    # We're good.
                    pass
                elif driver.current_url() == LOGIN_URL:
                    raise LoginError("GrDF sign in has failed, please check your username/password")
                else:
                    raise

            # Page 'votre consommation'
            driver.get(DATA_URL, "Go to 'Consommations' page")

            # Wait for the data page to load completely.
            time.sleep(self.__wait_time)

            # Eventually, close TokyWoky assistant which may hide the Download button.
            try:
                tokyWoky_close_button = driver.find_element_by_xpath("//div[@id='toky_container']/div/div", "TokyWoky assistant close button")
                tokyWoky_close_button.click()
            except Exception:
                # Do nothing, because the Pop up may not appear.
                pass

            # Select daily consumption
            daily_consumption_element = driver.find_element_by_xpath("//table[@id='_eConsoconsoDetaille_WAR_eConsoportlet_:idFormConsoDetaille:panelTypeGranularite1']/tbody/tr/td[3]/label", "Daily consumption button")
            daily_consumption_element.click()

            # Download file
            # xpath=//button[@id='_eConsoconsoDetaille_WAR_eConsoportlet_:idFormConsoDetaille:telechargerDonnees']/span
            download_button_element = driver.find_element_by_xpath("//button[@onclick=\"envoieGATelechargerConsoDetaille('particulier', 'jour_kwh');\"]/span", "Download button")
            download_button_element.click()

            # Timestamp of the data.
            data_timestamp = datetime.now().isoformat()

            # Wait a few for the download to complete
            time.sleep(self.__wait_time)

            # Load the XLSX file into the data structure
            file_list = glob.glob(data_file_path_pattern)

            for filename in file_list:

                Client.logger.debug(f"Loading Excel data file '{filename}'...")
                wb = load_workbook(filename=filename)
                ws = wb['Historique par jour']
                minRowNum = max(8, len(ws['B']) + 1 - self.__lastNRows) if self.__lastNRows > 0 else 8
                maxRowNum = len(ws['B'])
                for rownum in range(minRowNum, maxRowNum + 1):
                    row = {}
                    if ws.cell(column=2, row=rownum).value is not None:
                        row[PropertyNameEnum.DATE.value] = ws.cell(column=2, row=rownum).value
                        row[PropertyNameEnum.START_INDEX_M3.value] = ws.cell(column=3, row=rownum).value
                        row[PropertyNameEnum.END_INDEX_M3.value] = ws.cell(column=4, row=rownum).value
                        row[PropertyNameEnum.VOLUME_M3.value] = ws.cell(column=5, row=rownum).value
                        row[PropertyNameEnum.ENERGY_KWH.value] = ws.cell(column=6, row=rownum).value
                        row[PropertyNameEnum.CONVERTER_FACTOR.value] = ws.cell(column=7, row=rownum).value
                        row[PropertyNameEnum.LOCAL_TEMPERATURE.value] = ws.cell(column=8, row=rownum).value
                        row[PropertyNameEnum.TYPE.value] = ws.cell(column=9, row=rownum).value
                        row[PropertyNameEnum.TIMESTAMP.value] = data_timestamp
                        self.__data.append(row)
                wb.close()
                Client.logger.debug(f"Data read successfully between row #{minRowNum} and row #{maxRowNum}")

                os.remove(filename)

            Client.logger.debug("The data update terminates normally")
        except Exception:
            WebDriverWrapper.logger.error("An unexpected error occured while updating the data", exc_info=True)
        finally:
            # Quit the driver
            driver.quit()
